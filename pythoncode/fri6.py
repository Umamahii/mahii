from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1']=("InvoiceNo")
ws['B1']=("StockCode")
ws['C1']=("Description")
ws['D1']=("Quantity")
ws['E1']=("InvoiceDate")
import datetime
ws['E2'] = datetime.datetime.now()
ws['E3'] = datetime.datetime.now()
ws['E4'] = datetime.datetime.now()
ws['E5'] = datetime.datetime.now()
ws['F1']=("unitprice")
ws['G1']=("CustomerID")
ws['H1']=("Country")

wb.save("sample1.xlsx")